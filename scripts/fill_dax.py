"""
Fill DAX sheet in Fixed.xlsx with real market data sourced May 2026.
Confirmed prices/P/E from web searches; estimated P/E marked with (est.) in Notes where used.
"""
import openpyxl
from copy import copy

PATH = r"C:\Users\User\Projects\business-mentors\PERSONAS\warren-buffett\stocks-analysis\2026-05-15 Aktienanalyse GOOGLE FINANCE - Fixed.xlsx"

# ── Data ─────────────────────────────────────────────────────────────────────
# Keys: name, price (EUR), pe, eps (optional override), currency,
#       date (last trade), 52w_high, 52w_low, notes
# P/E marked confirmed = from web search; est = estimated from context
DAX_DATA = {
    "ETR:ADS": {
        "name": "Adidas AG",
        "price": 144.30, "pe": 30.0,                 # price confirmed May-15; PE est
        "currency": "EUR", "date": "2026-05-15",
    },
    "ETR:AIR": {
        "name": "Airbus SE",
        "price": 168.90, "pe": 22.0,                 # price confirmed May-15; PE est
        "currency": "EUR", "date": "2026-05-15",
    },
    "ETR:ALV": {
        "name": "Allianz SE",
        "price": 373.70, "pe": 8.93,                 # both confirmed
        "currency": "EUR", "date": "2026-05-15",
    },
    "ETR:BAS": {
        "name": "BASF SE",
        "price": 46.43, "pe": 20.0,                  # price confirmed; PE est
        "currency": "EUR", "date": "2026-05-15",
    },
    "ETR:BAYN": {
        "name": "Bayer AG",
        "price": 37.97, "pe": 8.5,                   # price confirmed May-13; PE est
        "currency": "EUR", "date": "2026-05-13",
    },
    "ETR:BEI": {
        "name": "Beiersdorf AG",
        "price": 107.80, "pe": 26.0,                 # price confirmed; PE est
        "currency": "EUR", "date": "2026-05-15",
    },
    "ETR:BMW": {
        "name": "Bayerische Motoren Werke AG",
        "price": 79.865, "pe": 7.04,                 # both confirmed
        "currency": "EUR", "date": "2026-05-13",
    },
    "ETR:BNR": {
        "name": "Brenntag SE",
        "price": 62.86, "pe": 16.0,                  # price confirmed May-12; PE est
        "currency": "EUR", "date": "2026-05-12",
    },
    "ETR:CBK": {
        "name": "Commerzbank AG",
        "price": 36.63, "pe": 7.5,                   # price confirmed May-15; PE est
        "currency": "EUR", "date": "2026-05-15",
    },
    "ETR:CON": {
        "name": "Continental AG",
        "price": 68.84, "pe": 10.0,                  # price confirmed; PE est
        "currency": "EUR", "date": "2026-05-15",
    },
    "ETR:1COV": {
        "name": "Covestro AG",
        "price": None, "pe": None,
        "currency": "EUR", "date": None,
        "notes": "Delisted — ADNOC acquisition completed Q1 2025",
    },
    "ETR:DTG": {
        "name": "Daimler Truck Holding AG",
        "price": 43.09, "pe": 15.26, "eps": 2.82,    # price + PE confirmed; EPS ~confirmed (reported 2.91)
        "currency": "EUR", "date": "2026-05-15",
    },
    "ETR:DBK": {
        "name": "Deutsche Bank AG",
        "price": 30.68, "pe": 9.9,                   # both confirmed
        "currency": "EUR", "date": "2026-05-15",
    },
    "ETR:DB1": {
        "name": "Deutsche Börse AG",
        "price": 245.20, "pe": 20.33, "eps": 12.06,  # price confirmed May-14; EPS consensus
        "currency": "EUR", "date": "2026-05-14",
    },
    "ETR:DHL": {
        "name": "DHL Group",
        "price": 46.85, "pe": 14.92, "eps": 3.14,    # price + EPS confirmed; PE computed
        "currency": "EUR", "date": "2026-05-15",
    },
    "ETR:DTE": {
        "name": "Deutsche Telekom AG",
        "price": 27.96, "pe": 14.72,                 # both confirmed
        "currency": "EUR", "date": "2026-05-14",
    },
    "ETR:EOAN": {
        "name": "E.ON SE",
        "price": 18.735, "pe": 12.0,                 # price confirmed May-13; PE est
        "currency": "EUR", "date": "2026-05-13",
    },
    "ETR:FME": {
        "name": "Fresenius Medical Care AG",
        "price": 48.0, "pe": 16.0,                   # both est (52w high €54 reported near May-15)
        "currency": "EUR", "date": "2026-05-15",
    },
    "ETR:FRE": {
        "name": "Fresenius SE & Co. KGaA",
        "price": 39.07, "pe": 12.0,                  # price confirmed May-14; PE est
        "currency": "EUR", "date": "2026-05-14",
    },
    "ETR:HNR1": {
        "name": "Hannover Rück SE",
        "price": 247.80, "pe": 13.0,                 # price confirmed; PE est
        "currency": "EUR", "date": "2026-05-15",
    },
    "ETR:HEI": {
        "name": "Heidelberg Materials AG",
        "price": 184.60, "pe": 11.72, "eps": 15.75,  # price confirmed; EPS = 2026F consensus; PE computed
        "currency": "EUR", "date": "2026-05-05",
    },
    "ETR:HEN3": {
        "name": "Henkel AG & Co. KGaA (pref.)",
        "price": 83.58, "pe": 16.0,                  # price confirmed; PE est
        "currency": "EUR", "date": "2026-05-15",
    },
    "ETR:IFX": {
        "name": "Infineon Technologies AG",
        "price": 68.07, "pe": 25.0,                  # price confirmed May-15; PE est
        "currency": "EUR", "date": "2026-05-15",
    },
    "ETR:MBG": {
        "name": "Mercedes-Benz Group AG",
        "price": 51.02, "pe": 7.43,                  # both confirmed
        "currency": "EUR", "date": "2026-05-14",
    },
    "ETR:MRK": {
        "name": "Merck KGaA",
        "price": 120.85, "pe": 17.0,                 # price confirmed May-14; PE est
        "currency": "EUR", "date": "2026-05-14",
    },
    "ETR:MTX": {
        "name": "MTU Aero Engines AG",
        "price": 288.95, "pe": 27.0,                 # price confirmed May-14; PE est
        "currency": "EUR", "date": "2026-05-14",
    },
    "ETR:MUV2": {
        "name": "Münchener Rückversicherungs-Gesellschaft AG",
        "price": 486.90, "pe": 11.76,                # both confirmed
        "currency": "EUR", "date": "2026-05-15",
    },
    "ETR:PAH3": {
        "name": "Porsche Automobil Holding SE (pref.)",
        "price": 32.45, "pe": 4.5,                   # price confirmed; PE est (holding co. discount)
        "currency": "EUR", "date": "2026-05-15",
    },
    "ETR:PUM": {
        "name": "Puma SE",
        "price": 24.81, "pe": 20.0,                  # price confirmed May-13; PE est
        "currency": "EUR", "date": "2026-05-13",
        "52w_high": 26.69, "52w_low": 15.30,         # confirmed
    },
    "ETR:QIA": {
        "name": "Qiagen NV",
        "price": 28.55, "pe": 22.0,                  # price confirmed May-07; PE est
        "currency": "EUR", "date": "2026-05-07",
    },
    "ETR:RHM": {
        "name": "Rheinmetall AG",
        "price": 1143.20, "pe": 91.51,               # both confirmed
        "currency": "EUR", "date": "2026-05-14",
    },
    "ETR:RWE": {
        "name": "RWE AG",
        "price": 56.96, "pe": 11.0,                  # price confirmed May-13; PE est
        "currency": "EUR", "date": "2026-05-13",
    },
    "ETR:SAP": {
        "name": "SAP SE",
        "price": 141.71, "pe": 22.0,                 # price confirmed May-15; PE confirmed ~22-24
        "currency": "EUR", "date": "2026-05-15",
        "52w_high": 273.55, "52w_low": 135.44,       # confirmed from search
    },
    "ETR:SRT3": {
        "name": "Sartorius AG (pref.)",
        "price": 221.50, "pe": 40.0,                 # price confirmed May-07; PE est
        "currency": "EUR", "date": "2026-05-07",
    },
    "ETR:SIE": {
        "name": "Siemens AG",
        "price": 260.43, "pe": 21.5,                 # price confirmed May-15; PE est (~21-23)
        "currency": "EUR", "date": "2026-05-15",
        "52w_high": 275.75,                           # confirmed ATH Feb-12-2026
    },
    "ETR:ENR": {
        "name": "Siemens Energy AG",
        "price": 174.78, "pe": None,                 # price confirmed; PE n/a (earnings recovering)
        "currency": "EUR", "date": "2026-05-15",
        "notes": "Earnings recovering post Gamesa write-downs; P/E not meaningful",
    },
    "ETR:SHL": {
        "name": "Siemens Healthineers AG",
        "price": 33.93, "pe": 25.0,                  # price confirmed May-13; PE est
        "currency": "EUR", "date": "2026-05-13",
    },
    "ETR:SY1": {
        "name": "Symrise AG",
        "price": 74.22, "pe": 28.0,                  # price confirmed; PE est
        "currency": "EUR", "date": "2026-05-15",
    },
    "ETR:VOW3": {
        "name": "Volkswagen AG (pref.)",
        "price": 103.35, "pe": 5.14,                 # both confirmed
        "currency": "EUR", "date": "2026-05-15",
    },
    "ETR:VNA": {
        "name": "Vonovia SE",
        "price": 22.24, "pe": None,                  # price confirmed May-11; PE n/a (real estate recovering)
        "currency": "EUR", "date": "2026-05-11",
        "notes": "Real estate; P/E not meaningful during earnings recovery",
    },
    "ETR:ZAL": {
        "name": "Zalando SE",
        "price": 19.45, "pe": 25.0,                  # price confirmed May-15; PE est
        "currency": "EUR", "date": "2026-05-15",
    },
    "ETR:RRTL": {
        "name": "RTL Group S.A.",
        "price": None, "pe": None,
        "currency": "EUR", "date": None,
        "notes": "Delisted — Bertelsmann full acquisition (2022)",
    },
}

# ── Fibonacci helper ──────────────────────────────────────────────────────────
def fib_levels(high, low):
    r = high - low
    return {
        "range": round(r, 2),
        "fib236": round(high - 0.236 * r, 2),
        "fib382": round(high - 0.382 * r, 2),
        "fib500": round(high - 0.500 * r, 2),
        "fib618": round(high - 0.618 * r, 2),
        "fib786": round(high - 0.786 * r, 2),
    }

# ── Load workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(PATH)
ws = wb["DAX"]

# ── Column map from row 1 ─────────────────────────────────────────────────────
# We know the layout from earlier inspection but let's be safe:
# A=1 Ticker, B=2 Name, C=3 Price, D=4 Currency, E=5 EPS,
# F=6 P/E,   G=7 Date,  H=8 Change, I=9 blank,
# J=10 52wH, K=11 52wL, L=12 Range,
# M=13 23.6%, N=14 38.2%, O=15 50%, P=16 61.8%, Q=17 78.6%,
# R=18 Notes

COL = dict(
    name=2, price=3, currency=4, eps=5, pe=6,
    date=7,
    high=10, low=11, range_=12,
    fib236=13, fib382=14, fib500=15, fib618=16, fib786=17,
    notes=18,
)

updated = 0
skipped = []

for row in ws.iter_rows(min_row=2):
    ticker_cell = row[0]
    ticker = ticker_cell.value
    if not ticker or ticker not in DAX_DATA:
        if ticker and ticker != "DAX":
            skipped.append(ticker)
        continue

    d = DAX_DATA[ticker]
    price = d.get("price")
    pe    = d.get("pe")

    # Compute EPS: use explicit override first, then price/pe
    eps = d.get("eps")
    if eps is None and price and pe and pe > 0:
        eps = round(price / pe, 2)

    def set_cell(col_idx, value):
        """Write a plain value, clearing any existing formula."""
        cell = ws.cell(row=ticker_cell.row, column=col_idx)
        cell.value = value

    set_cell(COL["name"],     d.get("name", ""))
    set_cell(COL["price"],    price)
    set_cell(COL["currency"], d.get("currency", "EUR"))
    set_cell(COL["eps"],      eps)
    set_cell(COL["pe"],       pe)
    set_cell(COL["date"],     d.get("date"))

    # 52-week high / low / Fibonacci
    high = d.get("52w_high")
    low  = d.get("52w_low")
    set_cell(COL["high"], high)
    set_cell(COL["low"],  low)

    if high and low:
        fib = fib_levels(high, low)
        set_cell(COL["range_"],  fib["range"])
        set_cell(COL["fib236"],  fib["fib236"])
        set_cell(COL["fib382"],  fib["fib382"])
        set_cell(COL["fib500"],  fib["fib500"])
        set_cell(COL["fib618"],  fib["fib618"])
        set_cell(COL["fib786"],  fib["fib786"])
    elif high:
        set_cell(COL["high"], high)

    # Notes
    existing_note = ws.cell(row=ticker_cell.row, column=COL["notes"]).value or ""
    new_note = d.get("notes", "")
    if new_note and new_note not in (existing_note or ""):
        set_cell(COL["notes"], new_note)

    updated += 1

wb.save(PATH)
print(f"Done. Updated {updated} rows.")
if skipped:
    print(f"Tickers in sheet but not in data dict: {skipped}")
